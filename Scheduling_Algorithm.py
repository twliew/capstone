import pandas as pd
import pulp
import openpyxl

#Load scored applicants file (output of Scoring_Algorithm.py)
input_file = pd.read_excel('scored_applicants.xlsx')

#Find grade column
grade_col_matches = [c for c in input_file.columns if 'grade' in c.lower()]
grade_col = grade_col_matches[0] if grade_col_matches else None
if grade_col:
    id_columns = ['Full Name', 'Grade', 'Email Address']
else:
    id_columns = ['Full Name', 'Email Address']
    print('WARNING: No grade column found in scored_applicants.xlsx')

#find weekly availability columns
scheduling_columns = [col for col in input_file.columns if '- Week' in col]

#rename long qualtrics column names to short Week 1, Week 2, etc.
week_rename     = {col: f'Week {i+1}' for i, col in enumerate(scheduling_columns)}
short_week_cols = [f'Week {i+1}' for i in range(len(scheduling_columns))]

#build pref_df (preference dataframe) from id columns + week columns only
pref_df = input_file[id_columns + scheduling_columns].copy()
pref_df = pref_df.rename(columns=week_rename)
pref_df.insert(0, 'Volunteer ID', pref_df.index + 1)
pref_map = {
    'I am available and prefer this week': 2,
    'I am available this week': 1,
    'I am not available': 0
}
#build availability_df (availability dataframe) which is a copy of pref_df but with values mapped to 1 for available and 0 for not available (used for hard constraints in the optimization)
avail_map = {
    'I am available and prefer this week': 1,
    'I am available this week': 1,
    'I am not available': 0
}
#copy before mapping so both start from raw text
availability_df = pref_df.copy()
for col in short_week_cols:
    pref_df[col]         = pref_df[col].map(pref_map)
    availability_df[col] = availability_df[col].map(avail_map)
weeks = list(range(1, len(short_week_cols) + 1))

#Get min and max volunteers per week from the Shiftly template
template_wb = openpyxl.load_workbook('Shiftly Template.xlsm', keep_vba=True)
template_ws = template_wb['Requirements Entry']
min_volunteers = []
row = 4
while True:
    val = template_ws[f'D{row}'].value
    if val is None:
        break
    min_volunteers.append(int(val))
    row += 1
print('min_volunteers:', min_volunteers)
max_volunteers = []
row = 4
while True:
    val = template_ws[f'E{row}'].value
    if val is None:
        break
    max_volunteers.append(int(val))
    row += 1
print('max_volunteers:', max_volunteers)

#Remove volunteers with zero availability across all weeks, as they cannot be scheduled and would cause issues in the optimization model
zero_avail = availability_df[availability_df[short_week_cols].sum(axis=1) == 0]
if not zero_avail.empty:
    print('WARNING: removed volunteers with no availability:')
    for name in zero_avail['Full Name'].tolist():
        print(f'  - {name}')
    availability_df = availability_df[availability_df[short_week_cols].sum(axis=1) > 0].reset_index(drop=True)
    pref_df         = pref_df[pref_df['Volunteer ID'].isin(availability_df['Volunteer ID'])].reset_index(drop=True)
volunteers = pref_df['Volunteer ID'].tolist()

#Prevent building a model if there are no volunteers with availability, as this would cause errors in the optimization solver
if not volunteers:
    print('ERROR: No volunteers have any availability. Cannot build a schedule.')
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    #Create an Excel file with diagnostics information about the issue
    wb_empty = Workbook()
    wd_empty = wb_empty.active
    wd_empty.title = 'Diagnostics'
    title_font_e   = Font(bold=True, size=13, name='Arial')
    error_font_e   = Font(name='Arial', color='9C0006')
    error_fill_e   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    section_font_e = Font(bold=True, size=11, name='Arial', color='FFFFFF')
    section_fill_e = PatternFill(start_color='5a5a5a', end_color='5a5a5a', fill_type='solid')
    wd_empty.cell(row=1, column=1, value='Schedule Diagnostics').font = title_font_e
    wd_empty.cell(row=2, column=1, value='ERROR: No volunteers had any availability. No schedule was produced.').font = error_font_e
    cur_e = 4
    hdr = wd_empty.cell(row=cur_e, column=1, value='Removed (No Availability)')
    hdr.font = section_font_e
    hdr.fill = section_fill_e
    wd_empty.cell(row=cur_e, column=2).fill = section_fill_e
    wd_empty.cell(row=cur_e, column=3).fill = section_fill_e
    cur_e += 1
    for name in zero_avail['Full Name'].tolist():
        wd_empty.cell(row=cur_e, column=1, value=name).font = error_font_e
        wd_empty.cell(row=cur_e, column=1).fill = error_fill_e
        wd_empty.cell(row=cur_e, column=2, value='Removed').font = error_font_e
        wd_empty.cell(row=cur_e, column=2).fill = error_fill_e
        wd_empty.cell(row=cur_e, column=3, value='Marked unavailable for all weeks').font = error_font_e
        wd_empty.cell(row=cur_e, column=3).fill = error_fill_e
        cur_e += 1
    wd_empty.column_dimensions['A'].width = 30
    wd_empty.column_dimensions['B'].width = 18
    wd_empty.column_dimensions['C'].width = 40
    wb_empty.save('Volunteer_Schedule.xlsx')
    print('Saved: Volunteer_Schedule.xlsx (diagnostics only — no schedule produced)')
    raise SystemExit(0)

#helper functions to get info about volunteers by their ID
def get_avail(vol_id, week):
    return availability_df.loc[availability_df['Volunteer ID'] == vol_id, f'Week {week}'].values[0]
def get_pref(vol_id, week):
    return pref_df.loc[pref_df['Volunteer ID'] == vol_id, f'Week {week}'].values[0]
def get_name(vol_id):
    return pref_df.loc[pref_df['Volunteer ID'] == vol_id, 'Full Name'].values[0]
def get_email(vol_id):
    return pref_df.loc[pref_df['Volunteer ID'] == vol_id, 'Email Address'].values[0]
def get_grade(vol_id):
    if 'Grade' not in pref_df.columns:
        return ''
    return pref_df.loc[pref_df['Volunteer ID'] == vol_id, 'Grade'].values[0]
def get_score(vol_id):
    row_position = pref_df.index[pref_df['Volunteer ID'] == vol_id][0]
    return input_file.at[row_position, 'Score']

#Goal Programming Optimization Model

#Minimization of: P1*(understaffing+overstaffing)+P2*unassigned_volunteers+P3*preference_misses-P4*availability_weighted_assignments-P5*score_weighted_assignments
model = pulp.LpProblem('Volunteer_Scheduling', pulp.LpMinimize)
x          = pulp.LpVariable.dicts('x', [(i, j) for i in volunteers for j in weeks], cat='Binary') #x[i,j] = 1 if volunteer i is assigned to week j, 0 otherwise
d_under    = pulp.LpVariable.dicts('d_under',    weeks, lowBound=0) #understaffing for each week (how many volunteers below the minimum)
d_over_max = pulp.LpVariable.dicts('d_over_max', weeks, lowBound=0) #overstaffing for each week (how many volunteers above the maximum)
unassigned = pulp.LpVariable.dicts('unassigned', volunteers, cat='Binary') #1 if volunteer is not assigned to any week, 0 if assigned to at least one week
pref_miss  = pulp.LpVariable.dicts('pref_miss', [(i, j) for i in volunteers for j in weeks], lowBound=0) #preference miss for each volunteer-week (1 if assigned to a week they are available for but do not prefer, 0 otherwise)
availability_counts = {i: sum(get_avail(i, j) for j in weeks) for i in volunteers}
max_avail = max(availability_counts.values())
min_avail = min(availability_counts.values())
'''
Weights for the optimization model: higher weight for understaffing than overstaffing, higher weight for unassigned volunteers, 
higher weight for preference misses, negative weight for assigning volunteers with higher availability and higher score to encourage the model to 
choose those volunteers when possible
'''
weights = {
    i: (max_avail - availability_counts[i]) / (max_avail - min_avail + 1)
    for i in volunteers
}

#Normalize scores to 0-1 range for weighting in the optimization model
scores = {i: get_score(i) for i in volunteers}
max_score = max(scores.values()) or 1
#Tie-breaking weights based on scores, normalized to 0-1 range
tie_weights = {i: scores[i] / max_score for i in volunteers}

#Constraints: staffing constraints for each week, each volunteer assigned to at least one week or marked unassigned, preference miss constraints for each volunteer-week, etc.
for j in weeks:
    assigned = pulp.lpSum(x[(i,j)] * get_avail(i,j) for i in volunteers)
    model += (assigned + d_under[j] >= min_volunteers[j-1],    f'Week_{j}_min')
    model += (assigned - d_over_max[j] <= max_volunteers[j-1], f'Week_{j}_max')
for i in volunteers:
    model += (
        pulp.lpSum(x[(i,j)] * get_avail(i,j) for j in weeks) + unassigned[i] >= 1,
        f'Vol_{i}_assigned'
    )
for i in volunteers:
    for j in weeks:
        not_preferred = 1 if (get_avail(i,j) == 1 and get_pref(i,j) != 2) else 0
        model += (pref_miss[(i,j)] >= x[(i,j)] * not_preferred, f'PrefMiss_{i}_{j}')
P1, P2, P3, P4, P5 = 1000, 100, 10, 1, 0.1 #weights for the optimization model
model += (
    P1 * pulp.lpSum(2*d_under[j] + d_over_max[j] for j in weeks)
    + P2 * pulp.lpSum(unassigned[i] for i in volunteers)
    + P3 * pulp.lpSum(pref_miss[(i,j)] for i in volunteers for j in weeks)
    - P4 * pulp.lpSum(weights[i]     * x[(i,j)] for i in volunteers for j in weeks)
    - P5 * pulp.lpSum(tie_weights[i] * x[(i,j)] for i in volunteers for j in weeks)
)
model.solve(pulp.PULP_CBC_CMD(msg=0)) #solve the optimization model
#Print results and diagnostics about the schedule
print('Status:', pulp.LpStatus[model.status])
print('\nWeekly staffing:')
for j in weeks:
    under    = pulp.value(d_under[j])
    over     = pulp.value(d_over_max[j])
    assigned = int(sum(pulp.value(x[(i,j)]) * get_avail(i,j) for i in volunteers))
    if under > 0.5:
        print(f'  Week {j}: SHORT by {int(round(under))} ({assigned} assigned, min={min_volunteers[j-1]})')
    elif over > 0.5:
        print(f'  Week {j}: OVER by {int(round(over))} ({assigned} assigned, max={max_volunteers[j-1]})')
    else:
        print(f'  Week {j}: OK ({assigned} assigned, min={min_volunteers[j-1]}, max={max_volunteers[j-1]})')
print('\nPlacements:')
for i in volunteers:
    if pulp.value(unassigned[i]) > 0.5:
        print(f'  {get_name(i)}: not placed')
    else:
        placed = [j for j in weeks if pulp.value(x[(i,j)]) * get_avail(i,j) > 0.5]
        print(f'  {get_name(i)}: week(s) {placed}')
        
#Export schedule to Excel with formatting
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb = Workbook()
#Schedule Sheet
ws = wb.active
ws.title = 'Schedule'
has_grade   = 'Grade' in pref_df.columns
cols_per_wk = 3 if has_grade else 2

col_groups = [
    tuple(get_column_letter(c + k) for k in range(cols_per_wk))
    for c in range(1, len(weeks) * cols_per_wk, cols_per_wk)
]

header_font  = Font(bold=True, name='Arial', color='FFFFFF')
header_fill  = PatternFill(start_color='5B3895', end_color='5B3895', fill_type='solid')
short_fill   = PatternFill(start_color='D81159', end_color='D81159', fill_type='solid')
over_fill    = PatternFill(start_color='F96D10', end_color='F96D10', fill_type='solid')
stripe_fill  = PatternFill(start_color='E0D8ED', end_color='E0D8ED', fill_type='solid')
white_fill   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
cell_font    = Font(name='Arial')
center_align = Alignment(horizontal='center')

for idx, cols in enumerate(col_groups):
    week_number = idx + 1
    name_col  = cols[0]
    email_col = cols[-1]
    row_fill  = stripe_fill if week_number % 2 == 1 else white_fill

    ws.merge_cells(f'{name_col}1:{email_col}1')
    cell           = ws[f'{name_col}1']
    cell.font      = header_font
    cell.alignment = center_align
    under = pulp.value(d_under[week_number])
    over  = pulp.value(d_over_max[week_number])
    if under > 0.5:
        cell.value = f'Week {week_number} SHORT {int(round(under))} (min={min_volunteers[week_number-1]})'
        cell.fill  = short_fill
    elif over > 0.5:
        cell.value = f'Week {week_number} OVER {int(round(over))} (max={max_volunteers[week_number-1]})'
        cell.fill  = over_fill
    else:
        cell.value = f'Week {week_number}'
        cell.fill  = header_fill

    row = 2
    for i in volunteers:
        if pulp.value(x[(i, week_number)]) * get_avail(i, week_number) > 0.5:
            ws[f'{cols[0]}{row}'].value     = get_name(i)
            ws[f'{cols[0]}{row}'].font      = cell_font
            ws[f'{cols[0]}{row}'].fill      = row_fill
            ws[f'{cols[0]}{row}'].alignment = center_align
            if has_grade:
                ws[f'{cols[1]}{row}'].value     = get_grade(i)
                ws[f'{cols[1]}{row}'].font      = cell_font
                ws[f'{cols[1]}{row}'].fill      = row_fill
                ws[f'{cols[1]}{row}'].alignment = center_align
            ws[f'{email_col}{row}'].value     = get_email(i)
            ws[f'{email_col}{row}'].font      = cell_font
            ws[f'{email_col}{row}'].fill      = row_fill
            ws[f'{email_col}{row}'].alignment = center_align
            row += 1

for col_idx in range(1, len(weeks) * cols_per_wk + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 22
#Diagnostics Sheet
wd = wb.create_sheet(title='Diagnostics')
title_font   = Font(bold=True, size=13, name='Arial')
section_font = Font(bold=True, size=11, name='Arial', color='FFFFFF')
section_fill = PatternFill(start_color='5B3895', end_color='5B3895', fill_type='solid')
ok_fill      = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill    = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
error_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ok_font      = Font(name='Arial', color='000000')
warn_font    = Font(name='Arial', color='000000')
error_font   = Font(name='Arial', color='000000')
left_align   = Alignment(horizontal='left')

def write_cell(sheet, row, col, value, font=None, fill=None, alignment=None):
    c = sheet.cell(row=row, column=col, value=value)
    if font:      c.font = font
    if fill:      c.fill = fill
    if alignment: c.alignment = alignment
    return c

write_cell(wd, 1, 1, 'Schedule Diagnostics', font=title_font, alignment=left_align)
wd.column_dimensions['A'].width = 30
wd.column_dimensions['B'].width = 40
wd.column_dimensions['C'].width = 40
cur = 3
#Weekly Staffing Summary
write_cell(wd, cur, 1, 'Weekly Staffing Summary', font=section_font, fill=section_fill, alignment=left_align)
write_cell(wd, cur, 2, '', fill=section_fill, alignment=left_align)
write_cell(wd, cur, 3, '', fill=section_fill, alignment=left_align)
cur += 1
write_cell(wd, cur, 1, 'Week',     font=Font(bold=True, name='Arial'), alignment=left_align)
write_cell(wd, cur, 2, 'Assigned', font=Font(bold=True, name='Arial'), alignment=left_align)
write_cell(wd, cur, 3, 'Status',   font=Font(bold=True, name='Arial'), alignment=left_align)
cur += 1
for j in weeks:
    under    = pulp.value(d_under[j])
    over     = pulp.value(d_over_max[j])
    assigned = int(sum(pulp.value(x[(i,j)]) * get_avail(i,j) for i in volunteers))
    if under > 0.5:
        f, fi, status = error_font, error_fill, f'SHORT by {int(round(under))} (min={min_volunteers[j-1]})'
    elif over > 0.5:
        f, fi, status = warn_font,  warn_fill,  f'OVER by {int(round(over))} (max={max_volunteers[j-1]})'
    else:
        f, fi, status = ok_font,    ok_fill,    f'OK (min={min_volunteers[j-1]}, max={max_volunteers[j-1]})'
    write_cell(wd, cur, 1, f'Week {j}', font=f, fill=fi, alignment=left_align)
    write_cell(wd, cur, 2, assigned,    font=f, fill=fi, alignment=left_align)
    write_cell(wd, cur, 3, status,      font=f, fill=fi, alignment=left_align)
    cur += 1
cur += 1
#Volunteer Placements
write_cell(wd, cur, 1, 'Volunteer Placements', font=section_font, fill=section_fill, alignment=left_align)
write_cell(wd, cur, 2, '', fill=section_fill, alignment=left_align)
write_cell(wd, cur, 3, '', fill=section_fill, alignment=left_align)
cur += 1
write_cell(wd, cur, 1, 'Volunteer',      font=Font(bold=True, name='Arial'), alignment=left_align)
write_cell(wd, cur, 2, 'Weeks Assigned', font=Font(bold=True, name='Arial'), alignment=left_align)
write_cell(wd, cur, 3, 'Status',         font=Font(bold=True, name='Arial'), alignment=left_align)
cur += 1
for i in volunteers:
    placed_weeks = [f'Week {j}' for j in weeks
                    if pulp.value(x[(i,j)]) * get_avail(i,j) > 0.5]
    if pulp.value(unassigned[i]) > 0.5:
        f, fi, status, weeks_str = error_font, error_fill, 'NOT PLACED', '—'
    else:
        f, fi, status, weeks_str = ok_font, ok_fill, 'Placed', ', '.join(placed_weeks)
    write_cell(wd, cur, 1, get_name(i), font=f, fill=fi, alignment=left_align)
    write_cell(wd, cur, 2, weeks_str,   font=f, fill=fi, alignment=left_align)
    write_cell(wd, cur, 3, status,      font=f, fill=fi, alignment=left_align)
    cur += 1

if not zero_avail.empty:
    cur += 1
    write_cell(wd, cur, 1, 'Removed (No Availability)', font=section_font, fill=section_fill, alignment=left_align)
    write_cell(wd, cur, 2, '', fill=section_fill, alignment=left_align)
    write_cell(wd, cur, 3, '', fill=section_fill, alignment=left_align)
    cur += 1
    for name in zero_avail['Full Name'].tolist():
        write_cell(wd, cur, 1, name,      font=error_font, fill=error_fill, alignment=left_align)
        write_cell(wd, cur, 2, 'Removed', font=error_font, fill=error_fill, alignment=left_align)
        write_cell(wd, cur, 3, 'Marked unavailable for all weeks', font=error_font, fill=error_fill, alignment=left_align)
        cur += 1
wb.save('Volunteer_Schedule.xlsx') #Save the output workbook
print('Saved: Volunteer_Schedule.xlsx')