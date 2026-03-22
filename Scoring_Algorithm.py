import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# Load Shiftly Template
shiftly_wb = openpyxl.load_workbook("Shiftly Template.xlsm", keep_vba=True)
req_entry_sheet = shiftly_wb['Requirements Entry']

# Import applicant data
df = pd.read_excel('applicants.xlsx')

# Generate question sheet names dynamically based on what's actually in the workbook
question_sheets = [s for s in shiftly_wb.sheetnames if s.startswith('Question_') and s.__contains__('Template') == False]
question_sheets = sorted(question_sheets, key=lambda s: int(s.split('_')[1]))
num_questions = len(question_sheets)

# ── Load question configs from each sheet ─────────────────────────────────────
question_configs = {}  # { question_text: {option: score, ...} }

for sheet_name in question_sheets:
    ws = shiftly_wb[sheet_name]

    question_text = ws['A2'].value

    option_score_map = {}
    row = 5
    while ws[f'A{row}'].value is not None and str(ws[f'A{row}'].value).strip() != '':
        option = str(ws[f'A{row}'].value).strip()
        score  = ws[f'B{row}'].value
        option_score_map[option] = score
        row += 1

    question_configs[question_text] = option_score_map

print(question_configs)

# Score cutoff
score_cutoff = req_entry_sheet['A7'].value or 0

# Clean column names
df.columns = df.columns.str.strip()
df_score = df.copy()

# ── Identify grade column ──────────────────────────────────────────────────────
grade_col_matches = [c for c in df.columns if 'grade' in c.lower()]
grade_col = grade_col_matches[0] if grade_col_matches else None
if grade_col:
    df_score['Grade'] = df_score[grade_col]
else:
    print("WARNING: No grade column found in applicants.xlsx")

# ── Identify weekly availability columns ──────────────────────────────────────
availability_cols = [c for c in df.columns if 'weekly availability' in c.lower()]
print(f"Found availability columns: {availability_cols}")

# ── Fully dynamic scoring (matches by column header name) ─────────────────────
dynamic_score_cols = []

for question_text, option_score_map in question_configs.items():
    score_col = f'{question_text} (score)'

    # Directly match by column name instead of value subset detection
    if question_text in df_score.columns:
        # Score = sum of scores for each selected option (handles multi-select)
        df_score[score_col] = df_score[question_text].apply(
            lambda cell: sum(
                option_score_map.get(val.strip(), 0)
                for val in str(cell).split(',')
            ) if pd.notna(cell) else 0
        )
        dynamic_score_cols.append(score_col)
    else:
        print(f"WARNING: No matching column found in CSV for question: '{question_text}'")

# Total score is purely the sum of all dynamic question scores
df_score['Score'] = df_score[dynamic_score_cols].sum(axis=1)

# Sort by score (highest to lowest)
df_score = df_score.sort_values(by='Score', ascending=False).reset_index(drop=True)

# ── Build final output column order ───────────────────────────────────────────
fixed_cols  = ['Full Name', 'Email Address', 'Grade']
answer_cols = list(question_configs.keys())
score_cols  = [f'{q} (score)' for q in question_configs.keys()]
tail_cols   = ['Score'] + availability_cols  # availability appended after Score

final_cols = fixed_cols + answer_cols + score_cols + tail_cols
final_cols = [c for c in final_cols if c in df_score.columns]

df_score_top = df_score[df_score["Score"] >= score_cutoff][final_cols]

# ── Save to Excel ──────────────────────────────────────────────────────────────
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.title = "Scored Applicants"

black_fill  = PatternFill("solid", fgColor="000000")
purple_fill = PatternFill("solid", fgColor="5B3895")
pink_fill   = PatternFill("solid", fgColor="D81159")
white_font  = Font(color="FFFFFF", bold=True)

answer_col_set      = set(answer_cols)
score_col_set       = set(score_cols) | {'Score'}
availability_col_set = set(availability_cols)

for col_idx, col_name in enumerate(df_score_top.columns, start=1):
    cell = output_ws.cell(row=1, column=col_idx, value=col_name)
    if col_name in answer_col_set:
        cell.fill = black_fill
        cell.font = white_font
    elif col_name in score_col_set:
        cell.fill = purple_fill
        cell.font = white_font
    elif col_name in availability_col_set:
        cell.fill = pink_fill
        cell.font = white_font

for row_idx, row in enumerate(df_score_top.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        output_ws.cell(row=row_idx, column=col_idx, value=value)

output_wb.save("scored_applicants.xlsx")
print("Saved scored_applicants.xlsx")